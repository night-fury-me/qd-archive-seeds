from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path

import pandas as pd  # type: ignore[import-untyped]
from openpyxl import load_workbook  # type: ignore[import-untyped]

from qdarchive_seeding.core.entities import AssetRecord, DatasetRecord
from qdarchive_seeding.infra.sinks._buffered import BufferedSink


@dataclass(slots=True)
class ExcelSink(BufferedSink):
    path: Path

    def __post_init__(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def _sheet_exists(self, sheet_name: str) -> bool:
        if not self.path.exists():
            return False
        wb = load_workbook(self.path, read_only=True)
        exists = sheet_name in wb.sheetnames
        wb.close()
        return exists

    def upsert_dataset(self, record: DatasetRecord) -> str:
        dataset_id = record.source_dataset_id or record.source_url
        keywords_str = ",".join(record.keywords) if record.keywords else ""
        persons_str = (
            ",".join(f"{p.name}:{p.role}" for p in record.persons) if record.persons else ""
        )
        row = {
            "id": dataset_id,
            "source_name": record.source_name,
            "source_dataset_id": record.source_dataset_id,
            "source_url": record.source_url,
            "title": record.title,
            "description": record.description,
            "doi": record.doi,
            "license": record.license,
            "year": record.year,
            "owner_name": record.owner_name,
            "owner_email": record.owner_email,
            "query_string": record.query_string,
            "repository_id": record.repository_id,
            "repository_url": record.repository_url,
            "version": record.version,
            "language": record.language,
            "upload_date": record.upload_date,
            "download_date": record.download_date,
            "download_repository_folder": record.download_repository_folder,
            "download_project_folder": record.download_project_folder,
            "download_version_folder": record.download_version_folder,
            "download_method": record.download_method,
            "is_harvested": record.is_harvested,
            "harvested_from": record.harvested_from,
            "keywords": keywords_str,
            "persons": persons_str,
        }
        return self._buffer_dataset(dataset_id, row)

    def upsert_asset(self, dataset_id: str, asset: AssetRecord) -> None:
        row = {
            "id": asset.asset_url,
            "dataset_id": dataset_id,
            "asset_url": asset.asset_url,
            "asset_type": asset.asset_type,
            "local_dir": asset.local_dir,
            "local_filename": asset.local_filename,
            "downloaded_at": asset.downloaded_at.isoformat() if asset.downloaded_at else None,
            "checksum_sha256": asset.checksum_sha256,
            "size_bytes": asset.size_bytes,
            "download_status": asset.download_status,
            "error_message": asset.error_message,
        }
        self._buffer_asset(asset.asset_url, row)

    def _flush_datasets(self) -> None:
        if not self._dataset_buffer:
            return
        new_df = pd.DataFrame(list(self._dataset_buffer.values()))
        if self._sheet_exists("datasets"):
            existing = pd.read_excel(self.path, sheet_name="datasets")
            # Drop rows that will be replaced by buffer entries
            existing = existing[~existing["id"].isin(self._dataset_buffer)]
            combined = pd.concat([existing, new_df], ignore_index=True)
        else:
            combined = new_df
        self._write_sheet("datasets", combined)
        self._dataset_buffer.clear()
        self._dataset_ops = 0

    def _flush_assets(self) -> None:
        if not self._asset_buffer:
            return
        new_df = pd.DataFrame(list(self._asset_buffer.values()))
        if self._sheet_exists("assets"):
            existing = pd.read_excel(self.path, sheet_name="assets")
            existing = existing[~existing["asset_url"].isin(self._asset_buffer)]
            combined = pd.concat([existing, new_df], ignore_index=True)
        else:
            combined = new_df
        self._write_sheet("assets", combined)
        self._asset_buffer.clear()
        self._asset_ops = 0

    def _write_sheet(self, sheet_name: str, df: pd.DataFrame) -> None:
        """Write a DataFrame to a specific sheet, preserving other sheets (atomic)."""
        if self.path.exists():
            all_sheets: dict[str, pd.DataFrame] = pd.read_excel(self.path, sheet_name=None)
            all_sheets[sheet_name] = df
        else:
            all_sheets = {sheet_name: df}

        tmp = self.path.with_suffix(".tmp")
        with pd.ExcelWriter(tmp) as writer:
            for name, sheet_df in all_sheets.items():
                sheet_df.to_excel(writer, sheet_name=name, index=False)
        os.replace(tmp, self.path)

    def get_existing_dataset_ids(self, repository_id: int) -> set[str]:
        return set()

    def get_pending_download_datasets(
        self, repository_id: int | None = None
    ) -> list[tuple[str, DatasetRecord, list[AssetRecord]]]:
        return []

    def get_file_statuses(self, dataset_id: str) -> dict[str, str]:
        return {}

